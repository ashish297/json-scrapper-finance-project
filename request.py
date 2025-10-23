import requests
import argparse
import json
import os
import re
from openpyxl import load_workbook
import pandas as pd
from api_client import fetch_officers_by_perm_id

cookies = {
    'x-cdn-account': 'eyJhbGciOiJSU0EtT0FFUCIsImVuYyI6IkExMjhDQkMtSFMyNTYiLCJraWQiOiIzOTI5YzU0NjEyMjk0Y2ZmYjM3ZDI5MDI1YjY2YTUyYiIsInppcCI6IkRFRiIsInR5cCI6IkpXVCIsImN0eSI6IkpXVCJ9.iMLektM7OUuPU6ktM6aIJMuz5ctXPvEf7O5c2MnFDnDh07sbREfUprNgmL3jCdCeM01W7yuHgp2baMqYuv2JXjaUkFmKRQcQN-_gw6DTqf3kIh-t7VqYuZLpacM1BFU_992ZMd-PqjxQU_pyo6LzRCdTXKzlswkHVC8jTkIyy_KD5-nLWVzHRQhxCEiIA_o_XTF6y5XIUBoEjqhGjgapXiq1ThuYLnohP68LuY8Gg0pu6B-wE5y45LSK6ofQNxj3yhF4SpKfWtqOhOL9j7yq1GGFrDreX7tBh-rCvCpKKureWQ3CN__54dcaKqS1Hzbf-eo4zhsJCV36L_zj23LWrw.ZxrAS-PD7EgzQjpfR3ZzDA.7Wm6MT-GzaohjOJReQMwgVxDS5QJU3vi6re8yUeyciKJexijfoxU8Qo-osQsc0FZSJIW8vO3K1H5jerCVZYxwIvDqNv2d-lHvqfC6ON08cMMJM_dan-PCymOXVNowS3GMzmvl_OjbedwGKbshdJxXjAoZaNfVph_NY_2Ejb1lnqUPBDQRhAM0wdHA-gNQ2-0vN87RVpexa4_cdypCILh68aDG0c87Gw6VMQ_Bq6YO75ZiwLkbTHvQnjKlt91_wfpLwUUaVkqaI7ThUMUYlXpq5wLd7tgFW5CvXqbvYhY8uIT8KY8R-VhgXbysS3_lntYBE9pN_Cuko0w7fBn-GbETP7isHgoGwLgLUSupBJMUnWuSJN62-DlnuPzNta0IAPuYOnJCx_yZZ9LvP1a2dwQI75q-UYr5qOkduvzI0fNwmQeBTqrO28iowWMex-4KO9VPmokNMFVakQ897iPTRQqxQJk9FxfU4IAbbH070qD8xig3XHYuatzZfRuO-MZ-GjE0DprsAGlx9_tC1bHtdng2bN2zWOQE9Xo_H93Q-1RFpXlAQhX4KjJzxrNp5qZ13b4x7VKUAz5SHFFF1nAsECkC6m377toXVhq3dyyuV1sgbsUPBQTLU2DC5vFJfYPoegI3lqvdeLUYf2Q68JYbLVgEP1sJ5XcVdGu-4U8OBZ8G-DHb0G_LX-ZFEXAaYcccUVry7b93qMli4oP-FsQrnamnTJNhE0Hq1Te6wdnfWn5uxFawmDUTZAXgBurjGJHhNxVJhIShMcnlsPf9HBqMzB3u3GNUBYDnxuFWatxrgpnwr9JJqO1Wh6BC1ITZZrjRg5DPts4bku03_3OdPWycH4QkKAAENl-6uafjvUYTkQOadn-Zf7qmTJSi1-fdIhKHup6tigak7hTNIRJ1o6cb_XFJ15wZve2HKLk23volfk0AUos5fQEnxkZlCB_dRuL7szxrXV4fOj2bWLu0nOWyy8wQcapZMDIebKHxXNCdLx8o72RekF4T8zEUbzLKhn1a7OKd0RNnZuKL-zOPDxeW03rUm2Eaa6ZZ9ahGo-5JJkhz1GU5i1oYQ-3XwbQphEf5mLSlJ3cp8gGCg3T2okL3INcUkjBsKOvED6RlFbJlFCyD5ZWHpeTxA-MJHG7sqK9QR7pwmAwqHHH5999c04C9dfQqO4E-AWlkZGRKqdUWKrVO5zQc-72smWIa_I28MMYv9rxZthPS8K09ueobzfCFp6zm7B_Tc9_PTpr2M4eAmzFg0zB-ftMlbRg4zqSHbvVegi1FvGMCeJ7KljsbJSP1qg6zEKc_3p6gXP593N0ffEHaYE_U6iL8fUVJNxjTrlggoZ5GwSuYIfBcRQomhYWfFQx_az35bcR5bx6eca1_xqUhHs68IymRs1Zd-i3u_dqmaYc78iu3OB9QL7SIUPUq480_DrZqYkmhalAMiDvxtmWcknR6lk86TUPYP_3mmFjPQMIiMkJ1IoCrj3qPo1VQ5hskJJxvz94V8dAIbdUKHdQfNXCIjJVmDGD2vVXuLdGspw5GDB9oZaL8F8gnfZkYX_vgwN410D2qoNnCOOqibXSblcLOpHCjBh1UE-eHOgrEwbOEFmvsqLDvNMU1g2JUJ_cdo_PsvncWl1H83aFoczknbjZlcC-X4E0CEBWfwgdqvsT0EgAgdDx5EFYKcqYQsmTO3b_up1R3Ka9Dk11Nz4F9L4-tplE22R71NUDqHXcp7-gHjeHZr2NzGLUVgd1X1wLOW9NPcTrlxCJsYIIcov9WB0.XSReZ6xHCyj3ekMsKplRtw',
    'x-sts-token': 'eyJhbGciOiJSUzI1NiIsImtpZCI6ImJlcGpHV0dkOW44WU9VQ1NwX3M3SXlRMmlKMFkzeWRFaHo1VDJJVlNqWTgiLCJ0eXAiOiJhdCtqd3QifQ.eyJhdWQiOiIzNDYzYTU5MjEzNDQ0NDI4YmVlNWVhZDg4ZmYxZDkwMDQ0NDFiNjU0IiwiZGF0YSI6IntcImNpcGhlcnRleHRcIjpcIjJoMVNpcXdGWTJGVkFRMlpPRUN2T2MxTnNFMGNoVlZUN0tGSWxGaThrNFIxQTdEeEFZTnVrd1VHU3lzemd5MFU4dGw3VE1LcXdQTVZQUTJ3NGFQbEhTRG56T0ZYeGN3Nmx6endJV25Xems1cG5DVC1QTjM2VHY0NUhaNEFQb3Q2cmRvdTdHd2ZYeVFUX1JiWVI1TG1nLVBaODJjTHdVVW1RaVZhNjJueEtKVFhvYlB0UVBlSlFJTHBuckc1eGxHRXpaSncxNW4xM2NtNGRqY3ZXV3VSbTBiUFMzakh4bXF6VXVxdlFRby1obGV4eDUzc3JRQkV0bUd5clBVbGt5REZGSjE1WFY4dS1fbV9IM1NHZUxaNWM0eVNZVnZqTk56cDdhNGs5V1RKS29OdjI2aTB3QjVzNUZ2SmsxV3VodWRkMS1LY0s5dGFwN0d1bkhSM1pVVXN4bDEzR0otV2l2dm5kTXBGMkRmTnBSTmd1REl0d1d0Y1R4S3JjQW94ZHVYUHJxVUljdnlSX1hDbkxIcHB3Yzg4b3JaNG9CWENhVjU5U1pRanJ0Q21Sd0pTX2V6NzdnQnRveGMzVHZ6Y2ZGWFl4XzRIRncyS2xVdzlnbDR3THc0clZmMnVoWUlud1Raa2h0N2kwR1lmMXhnaWFlMGRlSVN3TTZsdTNNQXYxYi1GSFJPU3NNc3Zrc3NXcDNFLThJSjRCTlV3bWQ5QV9qaEVXV3hIWTcxdWtDLTNWVmYyNjRyRjkwLVZVVzRBSjJ1azZ1ckZpTDh4eTIzODRlcThWa1NQV1RzaDg2b1IyaG51MFZqWHJfY1RaQW9GcnRMRjB4SkwxeHNJTGo0YnZQWXY1TExkWG83SmE4QzUxSDRmNktfMWlJU05Bc2VobnlaazBvbmpsMXB2Sjgwd3c1OTA2U00zbTd1OWYyWS1abEszQkgwN09KaVdnSzd2b1B0ZHF6N2wtNkFiclRMdkc5bVRvMkltSXpjczloYU9VSi10OFpUekw2NTY1dzVKMVNCaE4xbm9ybG9Zd2xRX3BtSTZOLUp6dnlFbXN3ck5Md3lxbTFnRVUxV3NpT1IwXCIsXCJlbmNyeXB0ZWRfa2V5XCI6XCJBUUlCQUhpU2VFcHVhS0ZYbHc5TUlkNXBEN01XN294dE9feWM4Sm5wZm9yeC03Yk9Td0dTaU5NdmZuMGlvb3A1Ui1XbURETE1BQUFBZmpCOEJna3Foa2lHOXcwQkJ3YWdiekJ0QWdFQU1HZ0dDU3FHU0liM0RRRUhBVEFlQmdsZ2hrZ0JaUU1FQVM0d0VRUU1Kbi1iYUpvbmVrSmIwcWFSQWdFUWdEdkx1bzRuaGN2U09vMGZEem9nRF95Q2Y5UWZtNl9OWUFpVzRXRXFyQVNlZ2Nicmt2ZVd1ZU1hQk83ZnVzclFsMzlvVkVMV0trTmFSYVpoUmdcIixcImhlYWRlclwiOntcImtpZFwiOlwiYXJuOmF3czprbXM6YXAtc291dGhlYXN0LTE6ODk4MDg0OTg0ODc3OmtleS9tcmstYzc5YzRjZjZiOTcwNDY0MWFkOGQ5MmY2MjQ0YzgzNTBcIn0sXCJpdlwiOlwic1UzZUl6ekVCSktyTFJyM1wiLFwicHJvdGVjdGVkXCI6XCJleUpoYkdjaU9pSkJWMU5mUlU1RFgxTkVTMTlCTWpVMklpd2laVzVqSWpvaVFUSTFOa2REVFNJc0lucHBjQ0k2SWtSRlJpSjlcIixcInRhZ1wiOlwiLVN1SjR3UFFJRXk3MDJfSVFJdVVUQVwifSIsImV4cCI6MTc2MDExNzg2NSwiaWF0IjoxNzYwMTE3MjY1LCJpc3MiOiJodHRwczovL2lkZW50aXR5LmNpYW0ucmVmaW5pdGl2LmNvbS9hcGkvaWRlbnRpdHkvc3RzX3Byb2QiLCJyczEiOiI0MDA1MzdmYWU0Yjc1MDYzODg3YTNmMjZlYjQ2NWQ3ZWZmZjRmN2M4In0.bWY80xdeYZmoJxBAbwxHfHLL958lkkqn5u8Mn1xiGYMQnPbGZ80ii-614L_GcBnQcqgvRnCxiis119wTocRT2CUwOXDcLpgSewvP6nYJgUrPRXxk7qWPyGLvQMnIWkHFgxVrD3C2wsy53oGJbnJ478W23hekjDAH_6jf31Q43hQNaaLRH59snTmN9e1Ute5eO8R5r8MEQEMLKE3ZeNHk0WrL8aIKaJSP3lVmdyiN6EEbnZCP9lKIxY5t3atPzxJOipAOQLkGFtEOT0CcEuUZEaaMKqnIOE8hVZe0DVPve2_B7ejPHDwVgawzW8YgCoFd6fM3crkixiqFRMenzCtWmQ',
    'hmds-token': 'AQIC5wM2LY4Sfcwt8OXwwUijkPbctWmwB0jBUsfbeKZmUIA%3D%40AAJTSQACMzAAAlNLABMzNjYwMjQzNDM0NjgxMTgwMDA0AAJTMQACMjI%3D%23',
    'x-login-domain': 'sts.identity.ciam.refinitiv.net',
    'sessionId': '7cf249ba-3104-4334-bae9-293d121b007e',
    'sharedSessionStartTime': '1760117268334',
    'userId': 'GESG1-237597',
    'BIGipServerHDCP-DATACLOUD-VIP-1080': '1219623178.45845.0000',
    'enableFeature': '',
    'disableFeature': '',
    'CSS': 'auto',
    'trace': '',
    '_dd_s': 'rum=0&expire=1760118179384&logs=1&id=584da007-39eb-401a-9b1d-3e115532b93f&created=1760117270544',
}

headers = {
    'accept': 'application/json, text/plain, */*',
    'accept-language': 'en-GB,en;q=0.5',
    'if-none-match': 'W/"5f0d9-HwAac+42O0Ae3sLT68GVzA"',
    'priority': 'u=1, i',
    'referer': 'https://workspace.refinitiv.com/Apps/OfficersDirectors/2.4.12/?view=detail&s=AAPL.OQ&st=RIC&corpId=corp_1760117278448',
    'sec-ch-ua': '"Brave";v="141", "Not?A_Brand";v="8", "Chromium";v="141"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Linux"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'sec-gpc': '1',
    'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36',
    # 'cookie': 'x-cdn-account=eyJhbGciOiJSU0EtT0FFUCIsImVuYyI6IkExMjhDQkMtSFMyNTYiLCJraWQiOiIzOTI5YzU0NjEyMjk0Y2ZmYjM3ZDI5MDI1YjY2YTUyYiIsInppcCI6IkRFRiIsInR5cCI6IkpXVCIsImN0eSI6IkpXVCJ9.iMLektM7OUuPU6ktM6aIJMuz5ctXPvEf7O5c2MnFDnDh07sbREfUprNgmL3jCdCeM01W7yuHgp2baMqYuv2JXjaUkFmKRQcQN-_gw6DTqf3kIh-t7VqYuZLpacM1BFU_992ZMd-PqjxQU_pyo6LzRCdTXKzlswkHVC8jTkIyy_KD5-nLWVzHRQhxCEiIA_o_XTF6y5XIUBoEjqhGjgapXiq1ThuYLnohP68LuY8Gg0pu6B-wE5y45LSK6ofQNxj3yhF4SpKfWtqOhOL9j7yq1GGFrDreX7tBh-rCvCpKKureWQ3CN__54dcaKqS1Hzbf-eo4zhsJCV36L_zj23LWrw.ZxrAS-PD7EgzQjpfR3ZzDA.7Wm6MT-GzaohjOJReQMwgVxDS5QJU3vi6re8yUeyciKJexijfoxU8Qo-osQsc0FZSJIW8vO3K1H5jerCVZYxwIvDqNv2d-lHvqfC6ON08cMMJM_dan-PCymOXVNowS3GMzmvl_OjbedwGKbshdJxXjAoZaNfVph_NY_2Ejb1lnqUPBDQRhAM0wdHA-gNQ2-0vN87RVpexa4_cdypCILh68aDG0c87Gw6VMQ_Bq6YO75ZiwLkbTHvQnjKlt91_wfpLwUUaVkqaI7ThUMUYlXpq5wLd7tgFW5CvXqbvYhY8uIT8KY8R-VhgXbysS3_lntYBE9pN_Cuko0w7fBn-GbETP7isHgoGwLgLUSupBJMUnWuSJN62-DlnuPzNta0IAPuYOnJCx_yZZ9LvP1a2dwQI75q-UYr5qOkduvzI0fNwmQeBTqrO28iowWMex-4KO9VPmokNMFVakQ897iPTRQqxQJk9FxfU4IAbbH070qD8xig3XHYuatzZfRuO-MZ-GjE0DprsAGlx9_tC1bHtdng2bN2zWOQE9Xo_H93Q-1RFpXlAQhX4KjJzxrNp5qZ13b4x7VKUAz5SHFFF1nAsECkC6m377toXVhq3dyyuV1sgbsUPBQTLU2DC5vFJfYPoegI3lqvdeLUYf2Q68JYbLVgEP1sJ5XcVdGu-4U8OBZ8G-DHb0G_LX-ZFEXAaYcccUVry7b93qMli4oP-FsQrnamnTJNhE0Hq1Te6wdnfWn5uxFawmDUTZAXgBurjGJHhNxVJhIShMcnlsPf9HBqMzB3u3GNUBYDnxuFWatxrgpnwr9JJqO1Wh6BC1ITZZrjRg5DPts4bku03_3OdPWycH4QkKAAENl-6uafjvUYTkQOadn-Zf7qmTJSi1-fdIhKHup6tigak7hTNIRJ1o6cb_XFJ15wZve2HKLk23volfk0AUos5fQEnxkZlCB_dRuL7szxrXV4fOj2bWLu0nOWyy8wQcapZMDIebKHxXNCdLx8o72RekF4T8zEUbzLKhn1a7OKd0RNnZuKL-zOPDxeW03rUm2Eaa6ZZ9ahGo-5JJkhz1GU5i1oYQ-3XwbQphEf5mLSlJ3cp8gGCg3T2okL3INcUkjBsKOvED6RlFbJlFCyD5ZWHpeTxA-MJHG7sqK9QR7pwmAwqHHH5999c04C9dfQqO4E-AWlkZGRKqdUWKrVO5zQc-72smWIa_I28MMYv9rxZthPS8K09ueobzfCFp6zm7B_Tc9_PTpr2M4eAmzFg0zB-ftMlbRg4zqSHbvVegi1FvGMCeJ7KljsbJSP1qg6zEKc_3p6gXP593N0ffEHaYE_U6iL8fUVJNxjTrlggoZ5GwSuYIfBcRQomhYWfFQx_az35bcR5bx6eca1_xqUhHs68IymRs1Zd-i3u_dqmaYc78iu3OB9QL7SIUPUq480_DrZqYkmhalAMiDvxtmWcknR6lk86TUPYP_3mmFjPQMIiMkJ1IoCrj3qPo1VQ5hskJJxvz94V8dAIbdUKHdQfNXCIjJVmDGD2vVXuLdGspw5GDB9oZaL8F8gnfZkYX_vgwN410D2qoNnCOOqibXSblcLOpHCjBh1UE-eHOgrEwbOEFmvsqLDvNMU1g2JUJ_cdo_PsvncWl1H83aFoczknbjZlcC-X4E0CEBWfwgdqvsT0EgAgdDx5EFYKcqYQsmTO3b_up1R3Ka9Dk11Nz4F9L4-tplE22R71NUDqHXcp7-gHjeHZr2NzGLUVgd1X1wLOW9NPcTrlxCJsYIIcov9WB0.XSReZ6xHCyj3ekMsKplRtw; x-sts-token=eyJhbGciOiJSUzI1NiIsImtpZCI6ImJlcGpHV0dkOW44WU9VQ1NwX3M3SXlRMmlKMFkzeWRFaHo1VDJJVlNqWTgiLCJ0eXAiOiJhdCtqd3QifQ.eyJhdWQiOiIzNDYzYTU5MjEzNDQ0NDI4YmVlNWVhZDg4ZmYxZDkwMDQ0NDFiNjU0IiwiZGF0YSI6IntcImNpcGhlcnRleHRcIjpcIjJoMVNpcXdGWTJGVkFRMlpPRUN2T2MxTnNFMGNoVlZUN0tGSWxGaThrNFIxQTdEeEFZTnVrd1VHU3lzemd5MFU4dGw3VE1LcXdQTVZQUTJ3NGFQbEhTRG56T0ZYeGN3Nmx6endJV25Xems1cG5DVC1QTjM2VHY0NUhaNEFQb3Q2cmRvdTdHd2ZYeVFUX1JiWVI1TG1nLVBaODJjTHdVVW1RaVZhNjJueEtKVFhvYlB0UVBlSlFJTHBuckc1eGxHRXpaSncxNW4xM2NtNGRqY3ZXV3VSbTBiUFMzakh4bXF6VXVxdlFRby1obGV4eDUzc3JRQkV0bUd5clBVbGt5REZGSjE1WFY4dS1fbV9IM1NHZUxaNWM0eVNZVnZqTk56cDdhNGs5V1RKS29OdjI2aTB3QjVzNUZ2SmsxV3VodWRkMS1LY0s5dGFwN0d1bkhSM1pVVXN4bDEzR0otV2l2dm5kTXBGMkRmTnBSTmd1REl0d1d0Y1R4S3JjQW94ZHVYUHJxVUljdnlSX1hDbkxIcHB3Yzg4b3JaNG9CWENhVjU5U1pRanJ0Q21Sd0pTX2V6NzdnQnRveGMzVHZ6Y2ZGWFl4XzRIRncyS2xVdzlnbDR3THc0clZmMnVoWUlud1Raa2h0N2kwR1lmMXhnaWFlMGRlSVN3TTZsdTNNQXYxYi1GSFJPU3NNc3Zrc3NXcDNFLThJSjRCTlV3bWQ5QV9qaEVXV3hIWTcxdWtDLTNWVmYyNjRyRjkwLVZVVzRBSjJ1azZ1ckZpTDh4eTIzODRlcThWa1NQV1RzaDg2b1IyaG51MFZqWHJfY1RaQW9GcnRMRjB4SkwxeHNJTGo0YnZQWXY1TExkWG83SmE4QzUxSDRmNktfMWlJU05Bc2VobnlaazBvbmpsMXB2Sjgwd3c1OTA2U00zbTd1OWYyWS1abEszQkgwN09KaVdnSzd2b1B0ZHF6N2wtNkFiclRMdkc5bVRvMkltSXpjczloYU9VSi10OFpUekw2NTY1dzVKMVNCaE4xbm9ybG9Zd2xRX3BtSTZOLUp6dnlFbXN3ck5Md3lxbTFnRVUxV3NpT1IwXCIsXCJlbmNyeXB0ZWRfa2V5XCI6XCJBUUlCQUhpU2VFcHVhS0ZYbHc5TUlkNXBEN01XN294dE9feWM4Sm5wZm9yeC03Yk9Td0dTaU5NdmZuMGlvb3A1Ui1XbURETE1BQUFBZmpCOEJna3Foa2lHOXcwQkJ3YWdiekJ0QWdFQU1HZ0dDU3FHU0liM0RRRUhBVEFlQmdsZ2hrZ0JaUU1FQVM0d0VRUU1Kbi1iYUpvbmVrSmIwcWFSQWdFUWdEdkx1bzRuaGN2U09vMGZEem9nRF95Q2Y5UWZtNl9OWUFpVzRXRXFyQVNlZ2Nicmt2ZVd1ZU1hQk83ZnVzclFsMzlvVkVMV0trTmFSYVpoUmdcIixcImhlYWRlclwiOntcImtpZFwiOlwiYXJuOmF3czprbXM6YXAtc291dGhlYXN0LTE6ODk4MDg0OTg0ODc3OmtleS9tcmstYzc5YzRjZjZiOTcwNDY0MWFkOGQ5MmY2MjQ0YzgzNTBcIn0sXCJpdlwiOlwic1UzZUl6ekVCSktyTFJyM1wiLFwicHJvdGVjdGVkXCI6XCJleUpoYkdjaU9pSkJWMU5mUlU1RFgxTkVTMTlCTWpVMklpd2laVzVqSWpvaVFUSTFOa2REVFNJc0lucHBjQ0k2SWtSRlJpSjlcIixcInRhZ1wiOlwiLVN1SjR3UFFJRXk3MDJfSVFJdVVUQVwifSIsImV4cCI6MTc2MDExNzg2NSwiaWF0IjoxNzYwMTE3MjY1LCJpc3MiOiJodHRwczovL2lkZW50aXR5LmNpYW0ucmVmaW5pdGl2LmNvbS9hcGkvaWRlbnRpdHkvc3RzX3Byb2QiLCJyczEiOiI0MDA1MzdmYWU0Yjc1MDYzODg3YTNmMjZlYjQ2NWQ3ZWZmZjRmN2M4In0.bWY80xdeYZmoJxBAbwxHfHLL958lkkqn5u8Mn1xiGYMQnPbGZ80ii-614L_GcBnQcqgvRnCxiis119wTocRT2CUwOXDcLpgSewvP6nYJgUrPRXxk7qWPyGLvQMnIWkHFgxVrD3C2wsy53oGJbnJ478W23hekjDAH_6jf31Q43hQNaaLRH59snTmN9e1Ute5eO8R5r8MEQEMLKE3ZeNHk0WrL8aIKaJSP3lVmdyiN6EEbnZCP9lKIxY5t3atPzxJOipAOQLkGFtEOT0CcEuUZEaaMKqnIOE8hVZe0DVPve2_B7ejPHDwVgawzW8YgCoFd6fM3crkixiqFRMenzCtWmQ; hmds-token=AQIC5wM2LY4Sfcwt8OXwwUijkPbctWmwB0jBUsfbeKZmUIA%3D%40AAJTSQACMzAAAlNLABMzNjYwMjQzNDM0NjgxMTgwMDA0AAJTMQACMjI%3D%23; x-login-domain=sts.identity.ciam.refinitiv.net; sessionId=7cf249ba-3104-4334-bae9-293d121b007e; sharedSessionStartTime=1760117268334; userId=GESG1-237597; BIGipServerHDCP-DATACLOUD-VIP-1080=1219623178.45845.0000; enableFeature=; disableFeature=; CSS=auto; trace=; _dd_s=rum=0&expire=1760118179384&logs=1&id=584da007-39eb-401a-9b1d-3e115532b93f&created=1760117270544',
}


def sanitize_filename_part(value: str) -> str:
    """Convert a string to a safe filename token: spaces -> '-', strip, and remove unsafe chars."""
    if value is None:
        return ""
    value = str(value).strip().replace(" ", "-")
    # Keep letters, numbers, dash and underscore; replace others with '-'
    value = re.sub(r"[^A-Za-z0-9_-]", "-", value)
    # Collapse multiple dashes
    value = re.sub(r"-+", "-", value)
    return value.strip("-")


def extract_org_info(json_data: dict) -> tuple[str, str] | tuple[None, None]:
    """Extract organisationName and OrgId from the nested response structure.

    Returns (organisation_name, org_id) or (None, None) if not available.
    """
    try:
        inner = json_data.get('response', {})
        inner = inner.get('response', inner)
        # Support both 'organisationName' and 'organizationName' just in case
        name = inner.get('organisationName') or inner.get('organizationName')
        org_id = inner.get('OrgId') or inner.get('orgId')
        if name and org_id:
            return str(name), str(org_id)
        return None, None
    except Exception:
        return None, None


def ensure_output_dir(path: str) -> bool:
    if os.path.isdir(path):
        return True
    if os.path.exists(path) and not os.path.isdir(path):
        print(f"ERROR: Output path '{path}' exists and is not a directory.")
        return False
    os.makedirs(path, exist_ok=True)
    return True


def fetch_and_save_for_perm_id(perm_id: str, output_dir: str) -> None:
    try:
        response = fetch_officers_by_perm_id(perm_id=perm_id, cookies=cookies, headers=headers, timeout_seconds=30)
    except Exception as e:
        print(f"Request failed for PermID {perm_id}: {e}")
        return

    print(f"PermID {perm_id}: HTTP {response.status_code}")
    content_type = response.headers.get('Content-Type', '')

    # Try to parse JSON; if it fails, save raw text with a fallback name
    data = None
    try:
        data = response.json()
    except ValueError:
        data = None

    org_name, org_id = (None, None)
    if data is not None:
        org_name, org_id = extract_org_info(data)

    if data is not None and org_name and org_id:
        safe_name = sanitize_filename_part(org_name)
        safe_id = sanitize_filename_part(org_id)
        filename = f"{safe_name}-{safe_id}.json"
        output_path = os.path.join(output_dir, filename)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"Saved JSON to {output_path}")
    elif data is not None:
        # JSON but missing expected fields — fall back to PermID in name
        filename = f"unknown-{sanitize_filename_part(perm_id)}.json"
        output_path = os.path.join(output_dir, filename)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"Saved JSON (missing org fields) to {output_path}")
    else:
        # Non-JSON response; save raw text
        filename = f"unknown-{sanitize_filename_part(perm_id)}.txt"
        output_path = os.path.join(output_dir, filename)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(response.text or '')
        print(f"Non-JSON response (Content-Type: {content_type}). Saved raw text to {output_path}")


def main(start_row: int | None = None) -> None:
    excel_path = '/home/ashish/Desktop/json-scrapper/main.xlsx'
    output_dir = '/home/ashish/Desktop/json-scrapper/JSON-DATA-ALL'

    if not ensure_output_dir(output_dir):
        return

    # Load workbook and find the 'Organization PermID' column
    print(f"Using Excel: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at {excel_path}")
        return
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    def normalize_header_name(value) -> str:
        s = str(value).strip().lower() if value is not None else ''
        # remove all non-alphanumeric characters
        return re.sub(r'[^a-z0-9]+', '', s)

    desired_variants = ['Organization PermID', 'PermID', 'Perm ID', 'OrganizationPermID', 'Org PermID', 'OrgPermID']
    desired_norms = {normalize_header_name(v) for v in desired_variants}

    # Search all sheets and the first 200 rows for a header row containing PermID
    found = False
    ws = None
    permid_col_idx = None
    header_row_index = None
    for sheet in wb.worksheets:
        max_scan_rows = min(200, sheet.max_row or 200)
        for r in range(1, max_scan_rows + 1):
            row_values = next(sheet.iter_rows(min_row=r, max_row=r, values_only=True))
            header_to_index = {str(h).strip(): idx for idx, h in enumerate(row_values) if h is not None and str(h).strip() != ''}
            normalized_to_index = {normalize_header_name(h): idx for h, idx in header_to_index.items()}
            # Try raw exact, case-insensitive exact, then normalized
            idx_candidate = None
            for v in desired_variants:
                if v in header_to_index:
                    idx_candidate = header_to_index[v]
                    break
            if idx_candidate is None:
                for k, idx in header_to_index.items():
                    if str(k).strip().lower() == 'permid':
                        idx_candidate = idx
                        break
            if idx_candidate is None:
                for norm, idx in normalized_to_index.items():
                    if norm in desired_norms:
                        idx_candidate = idx
                        break
            # As a final fallback, scan any cell whose normalized value equals 'permid'
            if idx_candidate is None:
                for idx, val in enumerate(row_values):
                    if normalize_header_name(val) == 'permid':
                        idx_candidate = idx
                        break
            if idx_candidate is not None:
                ws = sheet
                permid_col_idx = idx_candidate
                header_row_index = r
                found = True
                break
        if found:
            break

    if not found or ws is None or permid_col_idx is None or header_row_index is None:
        # Fallback: use pandas to read all sheets and scan for a column named PermID
        try:
            xls = pd.ExcelFile(excel_path)
            for sheet_name in xls.sheet_names:
                # Read the entire sheet (no row cap) so large files are processed fully
                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
                # attempt to find the header row by looking for any cell that matches desired headers
                header_row = None
                for r in range(len(df)):
                    row = df.iloc[r].astype(str).fillna('').tolist()
                    normalized_row = [normalize_header_name(v) for v in row]
                    if any(n in desired_norms for n in normalized_row):
                        header_row = r
                        break
                if header_row is None:
                    continue
                headers = df.iloc[header_row].astype(str).fillna('').tolist()
                header_to_index = {h.strip(): idx for idx, h in enumerate(headers) if h.strip()}
                normalized_to_index = {normalize_header_name(h): idx for h, idx in header_to_index.items()}
                # choose the first matching desired header by preference order
                permid_idx = None
                for desired in desired_variants:
                    norm = normalize_header_name(desired)
                    if norm in normalized_to_index:
                        permid_idx = normalized_to_index[norm]
                        break
                if permid_idx is None:
                    continue
                # Process rows after header
                start_data_idx = header_row + 1
                if start_row is not None:
                    # pandas dataframe index 0 corresponds to Excel row 1
                    start_data_idx = max(start_data_idx, max(1, int(start_row)) - 1)
                print(f"Sheet '{sheet_name}': header at row {header_row + 1}, start at row {start_data_idx + 1}, total rows {len(df)}")
                processed = 0
                for r in range(start_data_idx, len(df)):
                    row = df.iloc[r].tolist()
                    perm_id_cell = row[permid_idx] if permid_idx < len(row) else None
                    if pd.isna(perm_id_cell) or str(perm_id_cell).strip() == '':
                        continue
                    perm_id_str = str(perm_id_cell).strip()
                    fetch_and_save_for_perm_id(perm_id_str, output_dir)
                    processed += 1
                if processed == 0:
                    print("No rows processed. Is the start row beyond the available data?")
                # Mark as handled and return
                return
        except Exception as e:
            print(f"ERROR: Pandas fallback failed: {e}")
        # If still not found, print samples
        samples = []
        for sheet in wb.worksheets:
            row_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            samples.append(f"{sheet.title}: " + ", ".join([str(h) for h in row_values]))
        print("ERROR: 'PermID' column not found. Sample headers per sheet:\n" + "\n".join(samples))
        return

    # Determine starting data row (1-based Excel row numbers)
    start_data_row = header_row_index + 1
    if start_row is not None:
        start_row = max(1, int(start_row))
        start_data_row = max(start_data_row, start_row)
    print(f"Sheet '{ws.title}': header at row {header_row_index}, start at row {start_data_row}, max rows {ws.max_row}")

    # Iterate all rows starting from desired row
    processed = 0
    for row in ws.iter_rows(min_row=start_data_row, values_only=True):
        perm_id_cell = row[permid_col_idx] if permid_col_idx < len(row) else None
        if perm_id_cell is None:
            continue
        perm_id_str = str(perm_id_cell).strip()
        if not perm_id_str:
            continue
        fetch_and_save_for_perm_id(perm_id_str, output_dir)
        processed += 1
    if processed == 0:
        print("No rows processed. Try a smaller --start-row or verify the sheet has data after that row.")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Fetch and save Officers/Directors data by PermID from Excel.')
    parser.add_argument('-s', '--start-row', type=int, default=None,
                        help='1-based Excel row number to start processing from (inclusive). If this is before the header, processing starts at the first data row.')
    args = parser.parse_args()
    main(start_row=4162) # updated value for the start row



